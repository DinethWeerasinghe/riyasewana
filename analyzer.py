"""
analyzer.py - Car data analysis logic.
Accepts raw car data as a list of dicts (no file I/O).
"""

import csv
import io
import math
import pandas as pd


def _safe_float(val) -> float | None:
    """Return float or None for NaN/Inf (not JSON-serialisable)."""
    try:
        f = float(val)
        return None if (math.isnan(f) or math.isinf(f)) else round(f, 2)
    except (TypeError, ValueError):
        return None


def analyze_data(car_data: list) -> dict:
    """
    Analyse a list of car dicts (keys: year, price, full_title).
    Returns a JSON-safe dict with keys:
        summary       - overall statistics
        yearly_stats  - per-year statistics list
        price_buckets - price-range distribution
    All float values are sanitised (NaN → None).
    """
    # Filter valid rows
    valid = []
    for row in car_data:
        try:
            year = str(row.get('year', '')).strip()
            price = float(row.get('price', 0))
            if year and year != 'Unknown' and price > 0:
                valid.append({'year': int(year), 'price': price, 'full_title': row.get('full_title', '')})
        except (ValueError, TypeError):
            continue

    if not valid:
        return {'error': 'No valid data found in the scraped results.'}

    df = pd.DataFrame(valid)

    # Per-year statistics
    yearly = (
        df.groupby('year')
        .agg(
            avg_price=('price', 'mean'),
            count=('price', 'count'),
            min_price=('price', 'min'),
            max_price=('price', 'max'),
            median_price=('price', 'median'),
            std_price=('price', 'std'),  # NaN when count == 1
        )
        .reset_index()
        .sort_values('year')
    )

    yearly_list = [
        {
            'year':         int(r['year']),
            'count':        int(r['count']),
            'avg_price':    _safe_float(r['avg_price']),
            'median_price': _safe_float(r['median_price']),
            'min_price':    _safe_float(r['min_price']),
            'max_price':    _safe_float(r['max_price']),
            'std_price':    _safe_float(r['std_price']),
        }
        for _, r in yearly.iterrows()
    ]

    # Price buckets
    buckets = [
        ('<500k',   0,          500_000),
        ('500k-1M', 500_000,  1_000_000),
        ('1M-1.5M', 1_000_000, 1_500_000),
        ('1.5M-2M', 1_500_000, 2_000_000),
        ('2M-2.5M', 2_000_000, 2_500_000),
        ('2.5M-3M', 2_500_000, 3_000_000),
        ('>3M',     3_000_000, float('inf')),
    ]
    price_buckets = [
        {'label': label, 'count': int(((df['price'] >= lo) & (df['price'] < hi)).sum())}
        for label, lo, hi in buckets
    ]

    summary = {
        'total_cars':           len(valid),
        'year_min':             int(df['year'].min()),
        'year_max':             int(df['year'].max()),
        'overall_avg_price':    _safe_float(df['price'].mean()),
        'overall_min_price':    _safe_float(df['price'].min()),
        'overall_max_price':    _safe_float(df['price'].max()),
        'overall_median_price': _safe_float(df['price'].median()),
    }

    return {
        'summary':       summary,
        'yearly_stats':  yearly_list,
        'price_buckets': price_buckets,
    }


def build_download_excel(yearly_stats: list, raw_data: list) -> bytes:
    """
    Build a two-sheet Excel workbook:
      Sheet 1 'Yearly Summary'  — per-year price statistics
      Sheet 2 'All Listings'    — every individual car record
    Returns the workbook content as bytes (.xlsx).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter
    import io

    wb = Workbook()

    # ── Colour palette ────────────────────────────────────────────
    HEADER_FILL   = PatternFill('solid', fgColor='7C5CFC')   # purple accent
    HEADER_FONT   = Font(color='FFFFFF', bold=True, size=11)
    SECTION_FILL  = PatternFill('solid', fgColor='21262D')   # dark surface
    SECTION_FONT  = Font(color='A78BFA', bold=True, size=10)
    ALT_FILL      = PatternFill('solid', fgColor='161B22')   # slightly lighter
    PRICE_FMT     = '#,##0'
    PRICE_FMT_RS  = '"Rs."#,##0'

    def style_header_row(ws, row_num, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill   = HEADER_FILL
            cell.font   = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # ═══════════════════════════════════════════════════════════════
    # Sheet 1: Yearly Summary
    # ═══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Yearly Summary'

    headers1 = ['Year', 'Listings', 'Avg Price (Rs.)', 'Median Price (Rs.)',
                 'Min Price (Rs.)', 'Max Price (Rs.)', 'Std Dev (Rs.)']
    ws1.append(headers1)
    style_header_row(ws1, 1, len(headers1))
    ws1.row_dimensions[1].height = 22

    price_cols1 = [3, 4, 5, 6, 7]  # columns that contain prices

    for i, row in enumerate(yearly_stats, start=2):
        ws1.append([
            row.get('year'),
            row.get('count'),
            row.get('avg_price'),
            row.get('median_price'),
            row.get('min_price'),
            row.get('max_price'),
            row.get('std_price'),
        ])
        # Alternate row fill
        fill = ALT_FILL if i % 2 == 0 else SECTION_FILL
        for col in range(1, len(headers1) + 1):
            cell = ws1.cell(row=i, column=col)
            cell.fill = fill
            cell.font = Font(color='E6EDF3', size=10)
            cell.alignment = Alignment(horizontal='right' if col > 1 else 'center')
            if col in price_cols1 and cell.value is not None:
                cell.number_format = PRICE_FMT_RS

    auto_width(ws1)

    # ═══════════════════════════════════════════════════════════════
    # Sheet 2: All Listings
    # ═══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('All Listings')

    headers2 = ['Year', 'Price (Rs.)', 'Vehicle Title']
    ws2.append(headers2)
    style_header_row(ws2, 1, len(headers2))
    ws2.row_dimensions[1].height = 22

    for i, car in enumerate(raw_data, start=2):
        ws2.append([
            car.get('year'),
            car.get('price'),
            car.get('full_title'),
        ])
        fill = ALT_FILL if i % 2 == 0 else SECTION_FILL
        for col in range(1, len(headers2) + 1):
            cell = ws2.cell(row=i, column=col)
            cell.fill = fill
            cell.font = Font(color='E6EDF3', size=10)
            cell.alignment = Alignment(horizontal='left')
            if col == 2 and cell.value is not None:
                cell.number_format = PRICE_FMT_RS

    auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
